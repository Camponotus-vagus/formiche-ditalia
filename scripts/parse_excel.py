#!/usr/bin/env python3
"""Parse Excel 'Schede formiKEY' files to enrich genera/subfamilies/species JSON.

Reads descriptive sheets (one taxon per sheet, key-value pairs in columns A/B)
and merges into existing JSON data files produced by parse_nexus.py.

Excel structure (discovered from actual files):
  - Sottofamiglie: keys = Subfamily:, Description:, Identification:, Photos:
  - Generi: keys = Subfamily:, Genus:/Genera:, Description:, Identification:, Length:, Photos:
  - Specie: keys = Subfamily:, Genus:, Species:, Description:, Where to find it:, Length:, Photos:

Usage:
    python scripts/parse_excel.py
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "formiche-ditalia" / "src" / "data"
SCHEDE_DIR = PROJECT_ROOT / "TESI FORMICHE (Dropbox)" / "Schede formiKEY"


def load_json(filename: str) -> list[dict]:
    """Load a JSON array from the data directory."""
    path = DATA_DIR / filename
    if not path.exists():
        print(f"ERROR: {path} not found. Run parse_nexus.py first.")
        sys.exit(1)
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(filename: str, data: list[dict]):
    """Write a JSON array back to the data directory."""
    path = DATA_DIR / filename
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"  Updated {filename}: {len(data)} records")


def slugify(name: str) -> str:
    """Convert a taxon name to a URL-friendly slug."""
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def read_sheet_kv(ws) -> dict[str, str]:
    """Read a worksheet as key-value pairs from columns A/B.

    Keys are in column A (ending with ':'), values in column B.
    Rows with empty keys after the Photos row are treated as additional photo entries.
    """
    data = {}
    photos = []
    in_photos = False

    for row in ws.iter_rows(values_only=True):
        key_raw = str(row[0]).strip() if row[0] else ""
        val = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        if key_raw.endswith(":"):
            key = key_raw[:-1].strip()
            if key == "Photos":
                in_photos = True
                if val:
                    photos.append(val)
            else:
                in_photos = False
                if key and val:
                    data[key] = val
        elif in_photos and val:
            # Continuation rows under Photos: with empty key
            photos.append(val)

    if photos:
        data["Photos"] = photos

    return data


def extract_taxon_name(full_name: str) -> str:
    """Extract just the taxon name from 'Genus Author, Year' format.

    Examples:
        'Aphaenogaster Mayr, 1853' -> 'Aphaenogaster'
        'Leptanillinae Emery, 1910' -> 'Leptanillinae'
        'Aphaenogaster spinosa Emery, 1878' -> 'Aphaenogaster spinosa'
        'Aphaenogaster subterranea (Latreille, 1798)' -> 'Aphaenogaster subterranea'
    """
    # Remove trailing newlines/whitespace
    name = full_name.strip().rstrip("\n")

    # Try parenthetical author first: "Name (Author, Year)"
    m = re.match(r"^([\w\s]+?)\s*\(", name)
    if m:
        return m.group(1).strip()

    # Non-parenthetical author: "Name Author, Year" or "Name Author de Foo, Year"
    parts = re.split(r"\s+[A-Z][a-z]+(?:.*?\s+de\s+.*?)?,\s*\d{4}", name)
    if parts and parts[0].strip():
        return parts[0].strip()

    return name.strip()


def clean_text(text: str) -> str:
    """Clean up text values: normalize whitespace, remove NBSP."""
    if not text:
        return ""
    # Replace non-breaking spaces with regular spaces
    text = text.replace("\xa0", " ")
    # Normalize whitespace
    text = re.sub(r"\s+", " ", text).strip()
    return text


def enrich_subfamilies(subfamilies: list[dict]) -> int:
    """Enrich subfamilies from Schede sottofamiglie.xlsx. Returns match count."""
    xlsx_path = SCHEDE_DIR / "Sottofamiglie" / "Schede sottofamiglie.xlsx"
    if not xlsx_path.exists():
        print(f"  WARNING: {xlsx_path} not found, skipping subfamilies")
        return 0

    sf_by_id = {s["id"]: s for s in subfamilies}
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    matched = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kv = read_sheet_kv(ws)

        # Determine the subfamily name from the Subfamily: field or sheet name
        sf_name = sheet_name.strip()
        if "Subfamily" in kv:
            sf_name = extract_taxon_name(kv["Subfamily"])

        sf_id = slugify(sf_name)
        if sf_id not in sf_by_id:
            print(f"  WARNING: Subfamily '{sf_name}' (id={sf_id}) not in JSON, skipping")
            continue

        record = sf_by_id[sf_id]
        matched += 1

        # Description -> description_en (the sheets are in English)
        if "Description" in kv:
            record["description_en"] = clean_text(kv["Description"])

        # Identification -> diagnostic_characters (add field if needed)
        if "Identification" in kv:
            if "diagnostic_characters" not in record:
                record["diagnostic_characters"] = None
            record["diagnostic_characters"] = clean_text(kv["Identification"])

        # Photos
        if "Photos" in kv:
            if "photo_refs" not in record:
                record["photo_refs"] = []
            record["photo_refs"] = kv["Photos"]

        print(f"    Matched subfamily: {sf_name}")

    wb.close()
    return matched


def enrich_genera(genera: list[dict]) -> int:
    """Enrich genera from Schede generi.xlsx. Returns match count."""
    xlsx_path = SCHEDE_DIR / "Generi" / "Schede generi.xlsx"
    if not xlsx_path.exists():
        print(f"  WARNING: {xlsx_path} not found, skipping genera")
        return 0

    genera_by_id = {g["id"]: g for g in genera}
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    matched = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "esempio":
            continue

        ws = wb[sheet_name]
        kv = read_sheet_kv(ws)

        # Determine genus name from Genus: or Genera: field, or sheet name
        genus_name = sheet_name.strip()
        for key in ("Genus", "Genera"):
            if key in kv:
                genus_name = extract_taxon_name(kv[key])
                break

        genus_id = slugify(genus_name)
        if genus_id not in genera_by_id:
            print(f"  WARNING: Genus '{genus_name}' (id={genus_id}) not in JSON, skipping")
            continue

        record = genera_by_id[genus_id]
        matched += 1

        # Description -> description_en
        if "Description" in kv:
            record["description_en"] = clean_text(kv["Description"])

        # Identification -> diagnostic_characters
        if "Identification" in kv and clean_text(kv["Identification"]):
            record["diagnostic_characters"] = clean_text(kv["Identification"])

        # Length
        if "Length" in kv:
            if "body_length" not in record:
                record["body_length"] = None
            record["body_length"] = clean_text(kv["Length"])

        # Photos
        if "Photos" in kv:
            record["photo_refs"] = kv["Photos"]

        print(f"    Matched genus: {genus_name}")

    wb.close()
    return matched


def enrich_species(species: list[dict]) -> int:
    """Enrich species from Schede specie.xlsx. Returns match count."""
    xlsx_path = SCHEDE_DIR / "Specie" / "Schede specie.xlsx"
    if not xlsx_path.exists():
        print(f"  WARNING: {xlsx_path} not found, skipping species")
        return 0

    # Build lookup by ID and also by scientific_name for flexible matching
    species_by_id = {s["id"]: s for s in species}
    species_by_name = {}
    for s in species:
        name_lower = s["scientific_name"].lower().strip()
        species_by_name[name_lower] = s

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    matched = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "esempio":
            continue

        ws = wb[sheet_name]
        kv = read_sheet_kv(ws)

        # Determine species name from Species: field or sheet name
        sp_name = sheet_name.strip()
        if "Species" in kv:
            sp_name = extract_taxon_name(kv["Species"])

        sp_id = slugify(sp_name)

        # Try matching by ID first, then by name
        record = None
        if sp_id in species_by_id:
            record = species_by_id[sp_id]
        elif sp_name.lower() in species_by_name:
            record = species_by_name[sp_name.lower()]
        else:
            print(f"  WARNING: Species '{sp_name}' (id={sp_id}) not in JSON, skipping")
            continue

        matched += 1

        # Description -> description_en
        if "Description" in kv:
            record["description_en"] = clean_text(kv["Description"])

        # Where to find it -> habitat_notes
        if "Where to find it" in kv:
            record["habitat_notes"] = clean_text(kv["Where to find it"])

        # Length -> body_length
        if "Length" in kv:
            record["body_length"] = clean_text(kv["Length"])

        # Photos
        if "Photos" in kv:
            record["photo_refs"] = kv["Photos"]

        # Author/year from Species: field
        if "Species" in kv:
            full = kv["Species"].strip()
            # Try to extract author_year: everything after the species name
            taxon = extract_taxon_name(full)
            remainder = full[len(taxon):].strip()
            if remainder:
                # Clean up parenthetical authors: "(Latreille, 1798)" -> "Latreille, 1798"
                remainder = remainder.strip("()")
                if not record.get("author_year"):
                    record["author_year"] = remainder

        print(f"    Matched species: {sp_name}")

    wb.close()
    return matched


def main():
    print("=== Enriching JSON with Excel schede data ===\n")

    if not SCHEDE_DIR.exists():
        print(f"ERROR: {SCHEDE_DIR} not found.")
        sys.exit(1)

    genera = load_json("genera.json")
    subfamilies = load_json("subfamilies.json")
    species = load_json("species.json")

    # Enrich subfamilies
    print("Processing subfamilies...")
    sf_matched = enrich_subfamilies(subfamilies)
    print(f"  -> {sf_matched} subfamilies enriched\n")

    # Enrich genera
    print("Processing genera...")
    gen_matched = enrich_genera(genera)
    print(f"  -> {gen_matched} genera enriched\n")

    # Enrich species
    print("Processing species...")
    sp_matched = enrich_species(species)
    print(f"  -> {sp_matched} species enriched\n")

    # Save
    print("=== Saving enriched JSON ===")
    save_json("genera.json", genera)
    save_json("subfamilies.json", subfamilies)
    save_json("species.json", species)

    # Summary
    total = sf_matched + gen_matched + sp_matched
    print(f"\nDone! {total} records enriched total "
          f"({sf_matched} subfamilies, {gen_matched} genera, {sp_matched} species).")


if __name__ == "__main__":
    main()
